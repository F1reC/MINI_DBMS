# -*- coding: utf-8 -*

from glob import glob
from openpyxl import *
import  os
import re
from index import *
from prettytable import PrettyTable
import hashlib
import dbms_function
db_path = 'data/'
#view_path = 'view/'
user = ''

using_dbname = ''
using_db = Workbook()


def welcome():
    """
    欢迎界面/字符画
    :return:
    """
    print("""
          ##############################################
          
                    https://github.com/LANVNAL 
              _          _____  ____  __  __  _____ 
             | |        |  __ \|  _ \|  \/  |/ ____|
             | |  ______| |  | | |_) | \  / | (___  
             | | |______| |  | |  _ <| |\/| |\___ \ 
             | |____    | |__| | |_) | |  | |____) |
             |______|   |_____/|____/|_|  |_|_____/ 
                                                    
                    -> exit:退出 help:语法帮助 <-

          ##############################################
          """)


def help():
    """
    打印帮助信息
    :return:
    """
    print("""
    1.创建表：create database dbname
    2.创建数据库：create table tbname (id int PK null,user char[10] )
    3.删除：DELETE FROM table_nmae WHERE column_name = 'Value'
    4.更新：UPDATE table_name SET column1=value1,column2=value2,... WHERE some_column=some_value;
    5.插入：INSERT INTO table_name col1=val1,col2=val2&col3=val3,col4=val4
    6.查询：select a,b from table where c=x,d=x （与）
           select a,b from table where c=x|d=x（或）
           select a,b from table where c>x,d<x
           支持like，in，支持子查询
    7.权限：grant/revoke select on test_tb for testuser
    8.索引：creat view view_name as select xx from xx
    9.显示信息：help table/view/index
    """)

#使用数据库
def use_db(dbname):
    global using_dbname
    global using_db
    #数据库不存在
    if os.path.exists(db_path + dbname + '.xlsx'):
        if dbms_function.check_permission(user, dbname, 'use'):
            using_dbname = dbname
            print(dbname+"数据库已使用.")
            using_db = load_workbook(db_path+dbname+'.xlsx')
        else:
            print("你没有权限使用该数据库,请使用admin账户赋予权限.")
    else:
        print("数据库不存在")        

#显示所有数据库
def show_db():
    print("All database:")
    dbs = os.listdir(db_path)   #第二种方法，从保存数据库信息的库中查询
    for db in dbs:
        if '.DS' not in db and db != 'index':
            print("[*] " + db[:-5])

#创建数据库
def creat_db(dbname):
    dbpath = 'data/' + dbname + '.xlsx'
    database = Workbook()
    database.save(dbpath)
    dbms_function.create_tb_in_tbinfo(dbname)
    print(u"数据库创建操作执行成功")

def get_command():
    """
    从控制台获取命令
    :return: None
    """
    command = input("[👉]> ") if not using_dbname else input("[{}🚩]> ".format(using_dbname))
    #hcommand = command.lower()
    #print command
    return command.strip()

def Initialization():
    if not os.path.exists(db_path):
        os.mkdir(db_path)
    if not os.path.exists("data/table_information.xlsx"):
        Workbook().save("data/table_information.xlsx")
    if os.path.exists("data/system.xlsx"):
        print ("Initializating......")
    else:
        dbms_function.creat_db('system')
    db = load_workbook("data/system.xlsx")
    permission_tb_col = ['database char[50] pk unique','select char','insert char','delete char','update char']
    dbms_function.creat_table('permission', db, 'system',permission_tb_col)

def query(sql,tag=''):
    sql_word = sql.split(" ")
    if len(sql_word) < 2:
        print ("[!] Wrong query!")
        return
    operate = sql_word[0].lower()
    #使用数据库
    if operate == 'use':
        if sql_word[1] == 'database':
            try:
                use_db(sql_word[2])
            except:
                print ("[!]Error")
        else:
            print ("[!]Syntax Error.\neg:>use database dbname")
    #创建数据库、表、视图、索引
    elif operate == 'create':
        if sql_word[1] == 'database':
            try:
                creat_db(sql_word[2])
            except:
                print ("[!]Create Error")
        elif sql_word[1] == 'table':
            columns_list = re.findall('\((.*)\)', sql)[0].split(',')
            print (columns_list, using_dbname)
            try:
                dbms_function.creat_table(sql_word[2], using_db, using_dbname, columns_list)
            except:
                print ("[!]Error")
        elif sql_word[1] == 'view': #creat view test1 as select * from user
            viewname = sql_word[2]
            sql = ' '.join(sql_word[4:])
            dbms_function.view(viewname,sql)

        elif sql_word[1] == 'index':
            return
        else:
            print ("[!]Syntax Error.")
    #删除数据库或表
    elif operate == 'drop':
        if sql_word[1] == 'database':
            try:
                dbms_function.drop_db(sql_word[2])
            except:
                print ("[!]Error")
        if sql_word[1] == 'table':
            try:
                dbms_function.drop_table(sql_word[2],using_dbname,using_db)
            except:
                print ("[!]Error")
    #字段操作alter
    elif operate == 'alter':
        #添加字段
        if sql_word[2] == 'add':
            columns_list = re.findall('\((.*)\)', sql)[0].split(',')
            try:
                dbms_function.add_field(tbname = sql_word[1],columns_list=columns_list,using_dbname=using_dbname,using_db=using_db)
            except:
                print ("[!]Error")
        #删除字段
        elif sql_word[2] == 'drop':
            columns_list = re.findall('\((.*)\)', sql)[0].split(',')
            try:
                dbms_function.drop_field(tbname = sql_word[1],columns_list=columns_list,using_dbname=using_dbname,using_db=using_db)
            except:
                print ("[!]Error")
        #修改字段
        elif sql_word[2] == 'modify':
            columns_list = re.findall('\((.*)\)', sql)[0].split(',')
            try:
                dbms_function.modify_field(tbname = sql_word[1], alterFieldName = sql_word[3],columns_list=columns_list,using_dbname=using_dbname,using_db=using_db)
            except:
                print ("[!]Error")

    #选择操作select
    elif operate == 'select':
        pos = 0
        for i in range(len(sql_word)):
            if '(' in sql_word[i] and 'select' in sql_word[i]:
                pos = i
        if pos == 3:
            sql2 = sql_word[3][1:-1]
            query(sql2,tag='nesting')
            sql_word[3] = 'tmp'
            sql = ' '.join(sql_word)

        columns = sql_word[1]
        table_name = sql_word[3]
        if len(sql_word) > 4:
            #try:
            limit = sql_word[5].split()
            predicate = 'and'
            symbol = '='
            if ',' in sql_word[5]:
                limit = sql_word[5].split(',')
                predicate = 'and'
            elif '|' in sql_word[5]:
                limit = sql_word[5].split('|')
                predicate = 'or'
            elif '>' in sql_word[5]:
                #limit = sql_word[5].split()
                symbol = '>'
            elif '<' in sql_word[5]:
                #limit = sql_word[5].split()
                symbol = '<'
            elif len(sql_word) > 6:
                if sql_word[6] == 'in':
                    limit = [sql_word[5] + '=' + sql_word[7]]
                    predicate = 'in'
                if sql_word[6] == 'like':
                    limit = [sql_word[5] + '=' + sql_word[7]]
                    predicate = 'like'
            #except:
                #limit = [].append(sql_word[5])
            #print limit
            for i in range(len(limit)):
                limit[i] = limit[i].split(symbol)
            limit = dict(limit)
            return dbms_function.select(columns, table_name,using_dbname,using_db, limit, predicate=predicate, symbol=symbol, tag=tag)
        else:   #没where的情况
            return dbms_function.select(columns, table_name,using_dbname,using_db, tag=tag)
    #授予权限
    elif operate == 'grant':
        if user != 'admin':
            return  False
        dbms_function.set_permission(sql_word[5], sql_word[3], sql_word[1])
    #取消权限
    elif operate == 'revoke':
        if user != 'admin':
            return  False
        dbms_function.del_permission(sql_word[5], sql_word[3], sql_word[1])
    #插入数据
    elif operate == 'insert':   #INSERT INTO table_name col1=val1,col2=val2&col3=val3,col4=val4
        table_name = sql_word[2]
        """
        #INSERT INTO table_name (select x from xx)
        sql2 = re.findall('\((.*)\)')[0]
        query(sql2,tag='insert')
        """
        multiFlag = False

        columns_list = []
        if '&' in sql:
            multiFlag = True
            cols = sql_word[3].split('&')   #[{xx},{xx}] 多组
            for i in range(len(cols)):
                cols[i] = cols[i].split(',')
            for i in range(len(cols)):
                for j in range(len(cols[i])):
                    cols[i][j] = cols[i][j].split('=')
        else:
            cols = sql_word[3].split(',')
            for i in range(len(cols)):
                cols[i] = cols[i].split('=')
        dbms_function.insert_record(table_name,using_db,using_dbname,cols,multiFlag)
    #删除记录
    elif operate == 'delete':
        table_name = sql_word[2]
        if 'where' == sql_word[3]:
            if '&' in sql:
                cols = sql_word[4].split('&')
                for p in range(len(cols)):
                    col = cols[p]
                    condition_list = col.split(',')
            else:
                condition_list = sql_word[4].split(',')
            dbms_function.delete_record(table_name,using_db,using_dbname,condition_list)
        else:
            print ("[!]Syntax Error.")

    #修改记录
    elif operate == 'update':
        table_name = sql_word[1]
        #处理set后的=赋值部分
        if 'set' == sql_word[2]:
            multiFlag = False
            columns_list = []
            cols = sql_word[3].split(',')
            for i in range(len(cols)):
                cols[i] = cols[i].split('=')
        else:
            print ("[!]Syntax Error.")
        #处理where后的条件部分
        if 'where' == sql_word[4]:
            condition_list = sql_word[5].split(',')
        else:
            print ("[!]Syntax Error.")
        #调用函数update
        dbms_function.update_record(table_name,using_db,using_dbname,cols,condition_list,multiFlag)

            
    #帮助指令
    elif operate == 'help':
        if sql_word[1] == 'database':
            dbms_function.show_db()
        if sql_word[1] == 'table':
            usdbnm = using_dbname
            dbms_function.use_db('table_information')
            #若sql_word[2]存在，则指定表
            if len(sql_word) > 2 and sql_word[2] != '':
                tbname = sql_word[2]
                dbms_function.select('*',usdbnm,{'table':tbname})
            else:
                print ('[!]Syntax Error.\neg:>help table table_name')
        if sql_word[1] == 'view':
            view_name = sql_word[2]
            dbms_function.use_db('view')
            dbms_function.select('sql','sql',{'viewnamw':view_name})
        if sql_word[1] == 'index':
            print ("All Index:")
            indexs = os.listdir('data/index/')  # 第二种方法，从保存数据库信息的库中查询
            for index in indexs:
                if '.DS' not in index:
                    print ("[*] " + index[:-5])
    else:
        print ("[!]Syntax Error.")


def run():
    #Initialization()
    global user
    #welcome()
    user = dbms_function.login(user)
    while True:
        command = get_command()
        #print command
        if command == 'quit' or command == 'exit':
            print("[🍻] Thanks for using Mini DBMS. Bye~~")
            exit(0)
        elif command == 'help':
            help()
        else:
            query(command)

#若没有system和table_information库，则使用此方法创建创建
#if __name__ == '__main__':
    #Initialization()
    #run()
#登录
def userLogin(username,password,flagFirst,flagLogin):
    global user
    user,flagFirst,flagLogin = dbms_function.login(user,username,password,flagFirst,flagLogin)
    return flagFirst,flagLogin

def interpreter(command):
    if command == 'quit' or command == 'exit':
        print("[🍻] Thanks for using Mini DBMS. Bye~~")
        exit(0)
    elif command == 'help':
        help()
    else:
        query(command)